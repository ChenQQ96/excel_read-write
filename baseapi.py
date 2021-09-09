import pandas as pd
from openpyxl import load_workbook,Workbook 
import xlrd
import openpyxl
import os
import time

##xlrd用于读取公式计算结果，xls格式文件；新版本不可读取xls格式文件
##openpyxl用于读、写文件；无法读取由openpyxl保存的文件中，公式计算结果

#***********************************************************#
#打开文件
writer=openpyxl.load_workbook('计算表格0905.xlsx')

#指定工作表
sheet2=writer["Sheet2"]
nrows = sheet2.max_row # 获得行数
ncolumns = sheet2.max_column # 获得列数
print("目标sheet，行数：%d，列数：%d" %(nrows,ncolumns))
sheet5=writer["Sheet5"]
#***********************************************************#

#openpyxl读取参数
#压力P
P=[]
for i in range(11,18):
    P.append(sheet2.cell(i,2).value)
print(P)

#温度T
T=[]
for i in range(11,18):
    T.append(sheet2.cell(i,3).value)
print(T)
#气体流量Q
Q=[]
for i in range(4,24):
    Q.append(sheet5.cell(i,1).value)
print(Q)

#入口压力Pin
Pin=[]
for i in range(2,21):
    Pin.append(sheet5.cell(1,i).value)
print(Pin)

#出口压力Pout
Pout=[]
for i in range(2,21):
    Pout.append(sheet5.cell(2,i).value)
print(Pout)
print('---------------------------------------------')
# 将目标数值写入，并读取到指定位置
# 写入Q,Pin,Pout
#(0,20)


def xlsx2xls(file_name):
    """
    将xls文件另存为xlsx文件
    :param file_name: 要转换的文件路径
    :returns: new_excel_file_path 返回新的xlsx文件的路径
    """
    excel_file_path = file_name
    import win32com.client
    excel = win32com.client.DispatchEx('Excel.Application')
    wb = excel.Workbooks.Open(excel_file_path)
 
    new_excel_file_path = r'C:\Users\86187\Desktop\激波雾化器\计算表格0905.xls'
    if os.path.exists(new_excel_file_path):  # 先删掉新复制的文件
        os.remove(new_excel_file_path)
    wb.SaveAs(new_excel_file_path, FileFormat='52')# 51 表示的是xlsx格式；52 表示的是xls格式
    wb.Close()
    excel.Application.Quit()
    return new_excel_file_path

#(0,20)
for i in range(0,20):
    Q_cur=Q[i]
    #(0,19)
    for j in range(0,19):
        Pin_cur=Pin[j]
        Pout_cur=Pout[j]
        #写入参数
        sheet2.cell(1,4).value = Q_cur
        sheet2.cell(11,2).value=Pout_cur
        sheet2.cell(17,2).value=Pin_cur
        writer.save('计算表格0905.xlsx')

        print(Q_cur,Pout_cur,Pin_cur)
        new_excel_file_path=r'C:\Users\86187\Desktop\激波雾化器\计算表格0905.xls'
        xlsx2xls(r'C:\Users\86187\Desktop\激波雾化器\计算表格0905.xlsx')
        print('获取结果')
        #读取文件
        workbook=xlrd.open_workbook(new_excel_file_path)
        sheet2_name=workbook.sheet_by_name('Sheet2')
        d_cur=sheet2_name.cell(93,23).value
        print(d_cur)
        sheet5.cell(i+4,j+2).value=d_cur
        print('保存结果')

writer.save('计算表格0905.xlsx')
writer.close()